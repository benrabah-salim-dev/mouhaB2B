# b2b/views/importers.py
# -*- coding: utf-8 -*-
from __future__ import annotations

import io
import re
import difflib
import unicodedata
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from django.db import transaction
from django.shortcuts import get_object_or_404
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from openpyxl import load_workbook

from b2b.models import (
    AgenceVoyage,
    Dossier,
    Hotel,
    Vehicule,
    Chauffeur,
    FicheMouvement,
    FicheMouvementItem,
)

# ImportBatch peut être absent — on l'importe en souple
try:
    from b2b.models import ImportBatch, ImportBatchItem  # type: ignore
    HAS_BATCH = True
except Exception:
    ImportBatch = None  # type: ignore
    ImportBatchItem = None  # type: ignore
    HAS_BATCH = False

from b2b.views.helpers import (
    _ensure_same_agence_or_superadmin,
    _norm_header,
    _first_str,
    _parse_int_cell,
    _combine_datetime,
    normalize_flight_no,
    resolve_airports_and_type,
)

# ---------------------------------------------------------------------
# Lecture Excel robuste (buffer mémoire + CSV + multi-feuilles)
# ---------------------------------------------------------------------

def _norm_local(s: Any) -> str:
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return ""
    s = str(s)
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = s.lower().strip().replace("_", " ")
    s = re.sub(r"\s+", " ", s)
    return s


_EXPECTED_TOKENS = {
    "date","horaire","horaires","hora","provenance","org","origen","destination","dst","destino",
    "d/a","a/d","l/s","ls","depart/arriver","type","mouvement","n° vol","n vol","vuelo","flight","vol",
    "client/ to","client to","to","t.o.","tour operateur","tour opérateur","tour operador","hotel","hôtel",
    "ref","référence","reference","ntra.ref","ref t.o.","ref to","titulaire","tetulaire","titular","name",
    "holder","pax","passengers","adultes","adultos","enfants","niños","ninos","bb/gratuit","bebe","bebes",
    "observation","observations","coment","comentario","comments"
}

def _score_header_row(series: pd.Series) -> float:
    cells = [_norm_local(v) for v in series.tolist()]
    if not any(cells):
        return -1e9
    non_empty = sum(1 for c in cells if c)
    if non_empty < 2:
        return -1e9
    score = 0.0
    for c in cells:
        if not c:
            continue
        if any(tok in c for tok in _EXPECTED_TOKENS):
            score += 2.0
        if re.search(r"\d{3,}", c):
            score -= 0.25
    score += 0.15 * non_empty
    return score

def _tidy_df(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    df = df.dropna(axis=1, how="all").dropna(how="all")
    if df.empty:
        return pd.DataFrame()
    # colonnes
    fixed_cols, used = [], set()
    for i, c in enumerate(df.columns):
        nc = (c if isinstance(c, str) else "") or ""
        nc = nc.strip()
        if not nc or re.match(r"^unnamed", nc, re.I):
            nc = f"col_{i+1}"
        k = nc
        j = 2
        while k in used:
            k = f"{nc}_{j}"; j += 1
        used.add(k)
        fixed_cols.append(k)
    df.columns = fixed_cols
    # supprimer colonnes vides du début (tableau décalé)
    while df.shape[1] > 0:
        first_col = df.iloc[:, 0]
        ratio_nan = first_col.isna().mean()
        ratio_blank = (first_col.astype(str).str.strip() == "").mean()
        if max(ratio_nan, ratio_blank) >= 0.95:
            df = df.iloc[:, 1:]
        else:
            break
    # trim object
    for c in df.columns:
        if df[c].dtype == object:
            df[c] = df[c].apply(lambda x: x.strip() if isinstance(x, str) else x)
    df = df.dropna(how="all")
    return df if not df.empty else pd.DataFrame()

def smart_read_excel(file_like, max_header_scan: int = 200) -> pd.DataFrame:
    # charge en mémoire (évite curseur épuisé)
    if hasattr(file_like, "read"):
        raw_bytes = file_like.read()
    else:
        raw_bytes = file_like if isinstance(file_like, (bytes, bytearray)) else bytes(file_like)
    if not raw_bytes:
        return pd.DataFrame()
    bio = io.BytesIO(raw_bytes)
    filename = getattr(file_like, "name", "") or getattr(file_like, "filename", "") or ""

    # CSV rapide
    if filename.lower().endswith(".csv"):
        try:
            bio.seek(0)
            df = pd.read_csv(bio, dtype=str, keep_default_na=True)
            if df.empty:
                return pd.DataFrame()
            if any(str(c).lower().startswith("unnamed") for c in df.columns):
                bio.seek(0)
                raw = pd.read_csv(bio, header=None, dtype=str, keep_default_na=True)
                limit = min(max_header_scan, len(raw))
                best, header_idx = -1e9, 0
                for i in range(limit):
                    s = _score_header_row(raw.iloc[i])
                    if s > best:
                        best, header_idx = s, i
                headers = raw.iloc[header_idx].tolist()
                df = raw.iloc[header_idx+1:].copy()
                df.columns = headers
            return _tidy_df(df)
        except Exception:
            pass

    # Excel multi-feuilles
    def _parse_excel_from_memory() -> pd.DataFrame:
        try:
            bio.seek(0)
            xls = pd.ExcelFile(bio)
            sheets = xls.sheet_names or [0]
        except Exception:
            try:
                bio.seek(0)
                raw = pd.read_excel(bio, header=None, dtype=str)
                raw = raw.map(lambda x: x.strip() if isinstance(x, str) else x)
                limit = min(max_header_scan, len(raw))
                best, header_idx = -1e9, 0
                for i in range(limit):
                    s = _score_header_row(raw.iloc[i])
                    if s > best:
                        best, header_idx = s, i
                headers = raw.iloc[header_idx].tolist()
                df = raw.iloc[header_idx+1:].copy()
                df.columns = headers
                return _tidy_df(df)
            except Exception:
                return pd.DataFrame()

        for sh in sheets:
            try:
                raw = xls.parse(sh, header=None, dtype=str)
            except Exception:
                continue
            if raw is None or raw.empty:
                continue
            raw = raw.map(lambda x: x.strip() if isinstance(x, str) else x)
            limit = min(max_header_scan, len(raw))
            best, header_idx = -1e9, 0
            for i in range(limit):
                s = _score_header_row(raw.iloc[i])
                if s > best:
                    best, header_idx = s, i
            if best < -1e8:
                for i in range(min(200, len(raw))):
                    row = raw.iloc[i].astype(str).str.strip().replace("nan", "")
                    if row.astype(bool).any():
                        header_idx = i; break
            headers = raw.iloc[header_idx].tolist()
            df = raw.iloc[header_idx+1:].copy()
            df.columns = headers
            df = _tidy_df(df)
            if not df.empty:
                return df
        return pd.DataFrame()

    return _parse_excel_from_memory()


def _fuzzy_best_match(keywords: List[str], columns: List[str], min_ratio: float = 0.65) -> Optional[str]:
    if not columns:
        return None
    cols_norm = {c: _norm_header(c) for c in columns}
    best_c, best_r = None, 0.0
    for c, n in cols_norm.items():
        for k in keywords:
            r = difflib.SequenceMatcher(a=_norm_header(k), b=n).ratio()
            if r > best_r:
                best_c, best_r = c, r
    return best_c if best_r >= min_ratio else None


def _find_col(df: pd.DataFrame, *keyword_groups: List[str], prefer: Optional[str] = None) -> Optional[str]:
    norm_map: Dict[str, List[str]] = {}
    for c in df.columns:
        norm_map.setdefault(_norm_header(c), []).append(c)

    for group in keyword_groups:
        for k in group:
            k_norm = _norm_header(k)
            if k_norm in norm_map:
                cols = norm_map[k_norm]
                if prefer and prefer in cols:
                    return prefer
                return cols[0]

    for group in keyword_groups:
        for k in group:
            k_norm = _norm_header(k)
            candidates = [orig for norm, lst in norm_map.items() if k_norm in norm for orig in lst]
            if candidates:
                if prefer and prefer in candidates:
                    return prefer
                return candidates[0]
    return None


# ---------------------------------------------------------------------
# Import Dossiers
# ---------------------------------------------------------------------

class ImporterDossierAPIView(APIView):
    parser_classes = [MultiPartParser]
    permission_classes = [IsAuthenticated]

    def _find_cols_any(self, df: pd.DataFrame, *keywords: str) -> List[str]:
        want = {_norm_header(k) for k in keywords}
        results = []
        for col in df.columns:
            if any(w in _norm_header(col) for w in want):
                results.append(col)
        return results

    def _pick_hotel_col(self, df: pd.DataFrame) -> Optional[str]:
        candidates = [c for c in df.columns if "hotel" in _norm_header(c)]
        if not candidates:
            return None
        if len(candidates) == 1:
            return candidates[0]
        best_col, best_score = candidates[0], -1
        for col in candidates:
            ser = df[col].dropna().astype(str).head(50)
            score = sum(1 for v in ser if re.search(r"[A-Za-zÀ-ÿ]", v))
            if score > best_score:
                best_col, best_score = col, score
        return best_col

    def _extract_nom_reservation(self, df: pd.DataFrame, row: pd.Series) -> Optional[str]:
        cols = list(df.columns)
        norm_map = {c: _norm_header(c) for c in cols}

        def _good(col: str) -> bool:
            n = norm_map[col]
            if re.search(r"\b(to|tour|operat)\b", n):
                return False
            if "client" in n and "to" in n:
                return False
            return True

        g1_keys = ["titulaire","tetulaire","holder","lead","leadname","bookingname","titular"]
        g1 = [c for c in cols if any(k in norm_map[c] for k in g1_keys) and _good(c)]

        g2_keys = ["nomreservation","nomresa","reservation","groupe","group","booking","passager","paxnames"]
        g2 = [c for c in cols if any(k in norm_map[c] for k in g2_keys) and _good(c)]

        g3 = [c for c in cols if ("client" in norm_map[c]) and _good(c)]

        for group in (g1, g2, g3):
            for c in group:
                v = _first_str(row.get(c))
                if v:
                    s = re.sub(r"\s+", " ", v).strip()
                    if s and s.lower() not in {"nan", "none", "null", "-"}:
                        return s

        last_name_cols = self._find_cols_any(df, "nom","lastname","last_name","surname","apellidos")
        first_name_cols = self._find_cols_any(df, "prenom","firstname","first_name","givenname","nombre")
        ln = next((_first_str(row.get(c)) for c in last_name_cols if _first_str(row.get(c))), None)
        fn = next((_first_str(row.get(c)) for c in first_name_cols if _first_str(row.get(c))), None)
        combined = " ".join([fn or "", ln or ""]).strip()
        return combined or None

    def _collect_observations(self, df: pd.DataFrame, row: pd.Series) -> str:
        obs_exact_norms = {
            "observation","observations","observatio","observ","obs","remark","remarks","remarque","remarques",
            "note","notes","comment","comments","commentaire","commentaires","coment","coments","comentario","comentarios",
        }
        obs_num_re = re.compile(
            r"^(obs|observ|observation|observations|observatio|remark|remarks|remarque|remarques|"
            r"note|notes|comment|comments|commentaire|commentaires|coment|coments|comentario|comentarios)\d+$"
        )
        cols = list(df.columns)
        norm_map: Dict[str, List[str]] = {}
        for c in cols:
            norm_map.setdefault(_norm_header(c), []).append(c)
        obs_cols_set = set()
        for n, originals in norm_map.items():
            if n in obs_exact_norms:
                obs_cols_set.update(originals)
        for n, originals in norm_map.items():
            if obs_num_re.match(n):
                obs_cols_set.update(originals)
        if not obs_cols_set:
            for n, originals in norm_map.items():
                if any(k in n for k in obs_exact_norms):
                    obs_cols_set.update(originals)
        obs_cols = [c for c in cols if c in obs_cols_set]

        def _is_meaningful(val) -> bool:
            if val is None or (isinstance(val, float) and pd.isna(val)):
                return False
            s = str(val).strip()
            return bool(s) and s.lower() not in {"0", "0.0", "nan", "none", "null", "-"}

        def _clean_text(s: str) -> str:
            s = str(s).replace("\r", " ").replace("\n", " ")
            s = s.replace("T#", " ").replace("#", " ")
            s = re.sub(r"\s+", " ", s).strip()
            return s

        pieces, seen = [], set()
        for c in obs_cols:
            raw = row.get(c)
            if _is_meaningful(raw):
                txt = _clean_text(_first_str(raw) or "")
                if txt and txt not in seen:
                    seen.add(txt)
                    pieces.append(txt)
        return " | ".join(pieces)

    @transaction.atomic
    def post(self, request):
        fichier = request.FILES.get("file")
        agence_id = request.data.get("agence")
        if not fichier:
            return Response({"error": "Aucun fichier envoyé."}, status=400)
        if not agence_id:
            return Response({"error": "Agence requise."}, status=400)

        _ensure_same_agence_or_superadmin(request, int(agence_id))
        agence = get_object_or_404(AgenceVoyage, id=agence_id)

        try:
            df = smart_read_excel(fichier)
        except Exception as e:
            return Response({"error": "Erreur lecture Excel: {e}".format(e=e)}, status=400)
        if df.empty:
            return Response({"error": "Le fichier est vide."}, status=400)

        cols = list(df.columns)

        def choose_col(keywords, prefer=None):
            col = _find_col(df, keywords, prefer=prefer)
            return col or _fuzzy_best_match(keywords, cols, min_ratio=0.65)

        col_ref_to = choose_col(["Ref.T.O.","Ref TO","RefTO","RefTO.","Ref T.O.","Ref T O","Ref_T_O","Ref.TO"])
        col_ntra_ref = choose_col(["Ntra.Ref","NtraRef","Ntra Ref","Ntra"])
        col_ref_alt = choose_col(["Reference","Référence","Ref","REF","N° dossier","N dossier","N_DOSSIER"])
        col_ref = col_ref_to or col_ntra_ref or col_ref_alt

        col_day  = choose_col(["Dia","DATE","Date","Fecha","Jour","Data"])
        col_time = choose_col(["Hora","Horaires","Horaire","Heure","Time","Horas"])
        col_vol  = choose_col(["Vuelo","Vol","Flight","N° VOL","N VOL","Nº VOL","N° Vol","N°Vol","No Vol"])

        col_org = choose_col(["Org","Provenance","Orig","From","Origen"])
        col_dst = choose_col(["Dst","Destination","To","Destino"])

        col_ls = choose_col(["L/S","LS","D/A","A/D","DA","AD","Type Mouvement","Type","Mouvement",
                             "DEPART/ARRIVER","DEPART/ARRIVE","Depart/Arrivee","DEPART ARRIVEE"])

        col_city = choose_col(["Ciudad","Ville","City","Localite","Localité","Ciudad/Zone","Zone"])
        col_to   = choose_col(["T.O.","TO","Client TO","CLIENT/ TO","CLIENT TO","Client/ TO",
                               "Tour Operateur","Tour Opérateur","Tour Operador","Tour Operator"])
        col_hotel = self._pick_hotel_col(df)
        col_name  = choose_col(["Titular","Titulaire","TETULAIRE","Nom","Name","Holder","Client","Tetulaire"])

        col_pax = choose_col(["Pax","PAX","Passengers"])
        adult_cols = self._find_cols_any(df, "adulte","adultes","adults","adultos")
        child_cols = self._find_cols_any(df, "enfant","enfants","children","ninos","niños","nenes")
        baby_cols  = self._find_cols_any(df, "bb","bebe","bebes","bb/gratuit","infant","baby","bebesgratuit")

        dossiers_crees, dossiers_mis_a_jour = [], []
        lignes_ignorees, ui_rows = [], []
        created_ids, updated_ids = [], []

        def fmt_hhmm(dt):
            try:
                return dt.strftime("%H:%M") if dt else ""
            except Exception:
                return ""

        for idx, row in df.iterrows():
            try:
                # Référence obligatoire
                ref = None
                for c in [col_ref, col_ref_to, col_ntra_ref, col_ref_alt]:
                    if c and not ref:
                        ref = _first_str(row.get(c))
                if not ref:
                    lignes_ignorees.append({"ligne": idx + 2, "raison": "Référence manquante"})
                    continue
                ref = ref.strip()

                # Datetime
                day_val  = row.get(col_day) if col_day else None
                time_val = row.get(col_time) if col_time else None
                dt = _combine_datetime(day_val, time_val)

                # Orig/Dest + type
                org_val = row.get(col_org) if col_org else None
                dst_val = row.get(col_dst) if col_dst else None
                da_val  = row.get(col_ls)  if col_ls  else None
                org_iata, dst_iata, type_code, errs = resolve_airports_and_type(org_val, dst_val, da_val)

                # Vol, Ville, TO
                vol = normalize_flight_no(row.get(col_vol)) if col_vol else ""
                ville = _first_str(row.get(col_city)) if col_city else ""
                tour_op = _first_str(row.get(col_to)) if col_to else ""

                # PAX
                pax_raw = _first_str(row.get(col_pax)) if col_pax else None
                try:
                    pax = int(float(pax_raw)) if pax_raw is not None else 0
                except Exception:
                    pax = 0
                if pax <= 0:
                    ad = sum(_parse_int_cell(row.get(c)) for c in adult_cols)
                    ch = sum(_parse_int_cell(row.get(c)) for c in child_cols)
                    bb = sum(_parse_int_cell(row.get(c)) for c in baby_cols)
                    pax = max(pax, ad + ch + bb)

                # Nom de réservation
                nom_resa = self._extract_nom_reservation(df, row) or (_first_str(row.get(col_name)) if col_name else "")

                # Hôtel
                hotel_nom = _first_str(row.get(col_hotel)) if col_hotel else None
                hotel_obj = None
                if hotel_nom:
                    hotel_obj = Hotel.objects.filter(nom__iexact=hotel_nom).first()
                    if not hotel_obj:
                        hotel_obj = Hotel.objects.create(nom=hotel_nom)

                # A/D + vols
                heure_arrivee = heure_depart = None
                num_vol_arrivee = num_vol_retour = ""
                if type_code == "A":
                    heure_arrivee, num_vol_arrivee = dt, (vol or "")
                elif type_code == "D":
                    heure_depart,  num_vol_retour  = dt, (vol or "")
                else:
                    if org_iata and not dst_iata:
                        heure_depart,  num_vol_retour  = dt, (vol or "")
                    else:
                        heure_arrivee, num_vol_arrivee = dt, (vol or "")

                obs_joined = self._collect_observations(df, row)
                if errs:
                    obs_joined = (obs_joined + " | " if obs_joined else "") + "; ".join(errs)

                data = {
                    "agence": agence,
                    "ville": ville or "",
                    "aeroport_arrivee": (dst_iata or _first_str(dst_val) or "Aucun"),
                    "num_vol_arrivee": num_vol_arrivee or "",
                    "heure_arrivee": heure_arrivee,
                    "aeroport_depart": (org_iata or _first_str(org_val) or ""),
                    "heure_depart": heure_depart,
                    "num_vol_retour": num_vol_retour or "",
                    "hotel": hotel_obj,
                    "nombre_personnes_arrivee": pax if heure_arrivee else 0,
                    "nombre_personnes_retour": pax if heure_depart else 0,
                    "nom_reservation": nom_resa or "",
                    "tour_operateur": tour_op or "",
                    "observation": obs_joined or "",
                }

                obj, created = Dossier.objects.update_or_create(reference=ref, defaults=data)
                if created:
                    dossiers_crees.append(ref); created_ids.append(obj.id)
                else:
                    dossiers_mis_a_jour.append(ref); updated_ids.append(obj.id)

                # === créer / associer la fiche de mouvement
                fiche_type = "A" if data["heure_arrivee"] else ("D" if data["heure_depart"] else "")
                fiche_dt = (data["heure_arrivee"] or data["heure_depart"])
                fiche_date = fiche_dt.date() if fiche_dt else None
                fiche_airport = data["aeroport_arrivee"] if fiche_type == "A" else (data["aeroport_depart"] if fiche_type == "D" else "")

                if fiche_type and fiche_date:
                    fiche, _ = FicheMouvement.objects.get_or_create(
                        agence=agence, type=fiche_type, date=fiche_date, aeroport=fiche_airport or ""
                    )
                    FicheMouvementItem.objects.get_or_create(fiche=fiche, dossier=obj)

                # === UI row
                ui_type = "A" if obj.heure_arrivee else ("D" if obj.heure_depart else "")
                ui_date = ((obj.heure_arrivee or obj.heure_depart).date().isoformat()
                           if (obj.heure_arrivee or obj.heure_depart) else "")
                ui_airport = obj.aeroport_arrivee if ui_type == "A" else (obj.aeroport_depart if ui_type == "D" else "")
                ui_flight_no = obj.num_vol_arrivee if ui_type == "A" else (obj.num_vol_retour if ui_type == "D" else "")
                ui_flight_time = fmt_hhmm(obj.heure_arrivee if ui_type == "A" else obj.heure_depart)
                ui_pax_client = obj.nombre_personnes_arrivee or obj.nombre_personnes_retour or 0

                ui_rows.append({
                    "id": obj.id, "reference": obj.reference, "type": ui_type, "date": ui_date,
                    "aeroport": ui_airport, "flight_no": ui_flight_no, "flight_time": ui_flight_time,
                    "ville": obj.ville or "", "to": obj.tour_operateur or "", "_to": obj.tour_operateur or "",
                    "hotel": getattr(obj.hotel, "nom", None) or "",
                    "client_name": obj.nom_reservation or "", "nom_reservation": obj.nom_reservation or "",
                    "pax_client": ui_pax_client, "observation": obj.observation or "",
                    "aeroport_arrivee": obj.aeroport_arrivee, "num_vol_arrivee": obj.num_vol_arrivee,
                    "heure_arrivee": obj.heure_arrivee, "aeroport_depart": obj.aeroport_depart,
                    "heure_depart": obj.heure_depart, "num_vol_retour": obj.num_vol_retour,
                    "nombre_personnes_arrivee": obj.nombre_personnes_arrivee,
                    "nombre_personnes_retour": obj.nombre_personnes_retour,
                    "tour_operateur": obj.tour_operateur or "", "clients": obj.nom_reservation or "",
                })
            except Exception as e:
                lignes_ignorees.append({"ligne": idx + 2, "raison": f"{type(e).__name__}: {e}"})
                continue

        # === Batch d'import (optionnel)
        batch_id, batch_label = None, getattr(fichier, "name", "") or ""
        if HAS_BATCH:
            all_ids = list(dict.fromkeys(created_ids + updated_ids))
            if all_ids:
                batch = ImportBatch.objects.create(agence=agence, user=request.user, label=batch_label)  # type: ignore
                batch_id = str(batch.id)
                kept_ids = list(
                    Dossier.objects.filter(id__in=all_ids, agence=agence).values_list("id", flat=True)
                )
                ImportBatchItem.objects.bulk_create(  # type: ignore
                    [ImportBatchItem(batch=batch, dossier_id=i) for i in kept_ids], ignore_conflicts=True
                )
                
                
                            # ===== Mémoriser en session la sélection du DERNIER import =====
            # On ne met que les dossiers de l'import courant (créés ou mis à jour)
            imported_ids = [r["id"] for r in ui_rows if r.get("id")]
            request.session["last_import_dossier_ids"] = imported_ids
            request.session.modified = True


        return Response(
            {
                
                "message": "Import terminé",
                "agence": agence.id,
                "batch_id": batch_id,
                "batch_label": batch_label,
                "dossiers_crees": dossiers_crees,
                "dossiers_mis_a_jour": dossiers_mis_a_jour,
                "lignes_ignorees": lignes_ignorees,
                "dossiers": ui_rows,
                "created_count": len(created_ids),
                "updated_count": len(updated_ids),
                "total_importes": len(set(created_ids + updated_ids)),
            },
            
            status=200,
        )



# ---------------------------------------------------------------------
# Import Véhicules
# ---------------------------------------------------------------------

def _norm(s: Any) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    out = []
    for ch in s:
        out.append(ch if ch.isalnum() else "_")
    return "".join(out).strip("_").upper()


HEADER_ALIASES: Dict[str, List[str]] = {
    "IMMATRICULATION": ["IMMATRICULATION","IMMAT","PLAQUE","MATRICULE","REG","REG_NO","REGISTRATION","NUM_IMMATRICULATION"],
    "MARQUE": ["MARQUE","BRAND","MAKE"],
    "MODELE": ["MODELE","MODEL","MODELE_","MODElE","MODÈLE"],
    "TYPE": ["TYPE","CATEGORIE","CATEGORY","VEHICLE_TYPE"],
    "CAPACITE": ["CAPACITEE","CAPACITE","CAPACITE_","CAPACITY","SEATS","NB_PLACES","PLACES"],
    "ANNEE": ["ANNEE","ANNEE_","YEAR"],
}
REQUIRED_KEYS = ["IMMATRICULATION","MARQUE","MODELE"]

def _build_header_map(header_cells: List[Any]) -> Dict[str, int]:
    present = {_norm(v): idx for idx, v in enumerate(header_cells) if _norm(v)}
    mapping: Dict[str, int] = {}
    for logical, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            norm_alias = _norm(alias)
            if norm_alias in present:
                mapping[logical] = present[norm_alias]
                break
    return mapping

def _cell_val(cell) -> Any:
    if cell is None:
        return None
    v = cell.value
    if isinstance(v, str):
        return v.strip()
    return v


class ImporterVehiculesAPIView(APIView):
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request):
        file = request.FILES.get("file")
        agence_id = request.POST.get("agence") or request.data.get("agence")
        if not file:
            return Response({"error": "Aucun fichier reçu."}, status=400)
        if not agence_id:
            return Response({"error": "Paramètre 'agence' requis."}, status=400)

        _ensure_same_agence_or_superadmin(request, int(agence_id))
        agence = get_object_or_404(AgenceVoyage, id=agence_id)

        try:
            wb = load_workbook(file, data_only=True)
            ws = wb.active
        except Exception as e:
            return Response({"error": f"Fichier illisible ({e})."}, status=400)

        header_row_idx = None
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(20, ws.max_row))):
            labels = [(_cell_val(c) or "") for c in row]
            if any(str(v).strip() for v in labels):
                header_row_idx = i + 1
                header_cells = labels
                break
        if header_row_idx is None:
            return Response({"error": "Aucune ligne d'en-têtes détectée."}, status=400)

        header_map = _build_header_map(header_cells)
        missing = [k for k in REQUIRED_KEYS if k not in header_map]
        if missing:
            detected = ", ".join([f"{k}→col{header_map[k]+1}" for k in header_map.keys()])
            return Response(
                {"error": f"Colonnes manquantes dans le fichier: {', '.join(missing)}", "detected": detected},
                status=400,
            )

        created, updated, ignored = [], [], []
        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            row = [_cell_val(c) for c in ws[row_idx]]

            def colv(key, default=None):
                idx = header_map.get(key)
                if idx is None or idx >= len(row):
                    return default
                return row[idx]

            immat = (colv("IMMATRICULATION") or "").strip()
            if not immat:
                ignored.append({"ligne": row_idx, "raison": "Pas d'immatriculation"})
                continue

            marque = (colv("MARQUE") or "").strip()
            modele = (colv("MODELE") or "").strip()
            type_ = (colv("TYPE") or "").strip().lower() or "minibus"
            try:
                capacite = int(colv("CAPACITE") or 0)
            except Exception:
                capacite = 0
            try:
                annee = int(colv("ANNEE") or 0)
            except Exception:
                annee = 0

            if not marque or not modele:
                ignored.append({"ligne": row_idx, "raison": "MARQUE/MODELE manquant"})
                continue

            obj, was_created = Vehicule.objects.get_or_create(
                immatriculation=immat,
                defaults={
                    "type": (type_ if type_ in dict(Vehicule.TYPE_CHOICES) else "minibus"),
                    "marque": marque,
                    "model": modele,
                    "capacite": capacite,
                    "annee": annee or 0,
                    "agence": agence,
                },
            )
            if not was_created:
                changed = False
                if obj.agence_id != agence.id:
                    ignored.append({"ligne": row_idx, "raison": f"Immat déjà utilisée par une autre agence ({obj.agence_id})."})
                    continue
                if marque and obj.marque != marque:
                    obj.marque = marque; changed = True
                if modele and obj.model != modele:
                    obj.model = modele; changed = True
                if type_ and type_ in dict(Vehicule.TYPE_CHOICES) and obj.type != type_:
                    obj.type = type_; changed = True
                if capacite and obj.capacite != capacite:
                    obj.capacite = capacite; changed = True
                if annee and obj.annee != annee:
                    obj.annee = annee; changed = True
                if changed:
                    obj.save()
                    updated.append({"id": obj.id, "immatriculation": obj.immatriculation})
                else:
                    ignored.append({"ligne": row_idx, "raison": "Aucune modification."})
            else:
                created.append({"id": obj.id, "immatriculation": obj.immatriculation})

        return Response({"vehicules_crees": created, "vehicules_mis_a_jour": updated, "lignes_ignorees": ignored}, status=200)


# ---------------------------------------------------------------------
# Import Chauffeurs
# ---------------------------------------------------------------------

class ImporterChauffeursAPIView(APIView):
    parser_classes = [MultiPartParser]
    permission_classes = [IsAuthenticated]

    HEADERS = {
        "nom": ["NOM", "Nom", "Last name", "Apellido"],
        "prenom": ["PRENOM", "Prénom", "First name", "Nombre"],
        "cin": ["CIN", "N° CIN", "C.I.N", "ID", "Identité"],
    }

    def _find_col(self, df, candidates):
        for c in candidates:
            if c in df.columns:
                return c
        lowered = {str(col).strip().lower(): col for col in df.columns}
        for c in candidates:
            key = str(c).strip().lower()
            if key in lowered:
                return lowered[key]
        return None

    def _clean_str(self, val):
        if pd.isna(val) or val is None:
            return ""
        return str(val).strip()

    def post(self, request, *args, **kwargs):
        fichier = request.FILES.get("file")
        agence_id = request.data.get("agence")
        if not fichier:
            return Response({"error": "Aucun fichier envoyé."}, status=400)
        if not agence_id:
            return Response({"error": "Aucune agence spécifiée."}, status=400)
        _ensure_same_agence_or_superadmin(request, int(agence_id))
        agence = get_object_or_404(AgenceVoyage, id=agence_id)

        try:
            df = pd.read_excel(fichier)
        except Exception as e:
            return Response({"error": f"Erreur lecture fichier Excel: {e}"}, status=400)

        col_nom = self._find_col(df, self.HEADERS["nom"])
        col_prenom = self._find_col(df, self.HEADERS["prenom"])
        col_cin = self._find_col(df, self.HEADERS["cin"])

        if not col_nom:
            return Response({"error": "Colonne NOM manquante."}, status=400)

        created, updated, ignored = [], [], []
        for idx, row in df.iterrows():
            nom = self._clean_str(row.get(col_nom))
            prenom = self._clean_str(row.get(col_prenom)) if col_prenom else ""
            cin = self._clean_str(row.get(col_cin)) if col_cin else ""

            if not nom:
                ignored.append({"ligne": idx + 2, "raison": "Nom manquant"})
                continue

            obj, was_created = Chauffeur.objects.update_or_create(
                agence=agence,
                nom=nom,
                prenom=prenom or "",
                defaults={"cin": cin or "", "agence": agence, "nom": nom, "prenom": prenom or ""},
            )
            (created if was_created else updated).append(f"{nom} {prenom}".strip())

        return Response(
            {
                "message": "Import chauffeurs terminé",
                "agence": agence.id,
                "chauffeurs_crees": created,
                "chauffeurs_mis_a_jour": updated,
                "lignes_ignorees": ignored,
                "resume": {
                    "crees": len(created),
                    "mis_a_jour": len(updated),
                    "ignores": len(ignored),
                    "total_lues": int(df.shape[0]),
                },
            },
            status=200,
        )
